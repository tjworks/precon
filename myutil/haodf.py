#!/usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import codecs
import requests
import traceback

DataFile = 'haodf-telephone.xlsx'
 
#configBook = load_workbook(filename=ConfigFile)
#dataBook = load_workbook(filename=DataFile)
"""
def listAllCases():
    datasheet = dataBook.worksheets[0]
    rowCount = datasheet.get_highest_row()
    cases = []
    for i in range(rowCount):
        if(i<=3): continue
        case = {}
        case['judge'] = datasheet.cell(row=i, column = I).value or ''
        case['case_id'] = datasheet.cell(row=i, column = A).value or ''
        case['docket_number'] = datasheet.cell(row=i, column = D).value or ''
        cases.append(case)
    return cases
"""


copyed = u'医院 科室 '
b = None
def download():
    base_url = "http://400.haodf.com/index/search?diseasename=&province=&facultyname=&hosfaculty=&hospitalname=&nowpage="
    txt =  '' 
    for i in range(500):
        pageno  = i+1
        #if( pageno == 2): break
        print "Downloading page %d" %(pageno)
        url = "%s%s" %(base_url, (pageno) )
        r= requests.get(url)
        print "Done"
        txt = r.text
        with codecs.open("files/%s.html" % pageno, mode='w', encoding='utf-8') as f:
            f.write(txt)
        with codecs.open("files/%s.html" %pageno, mode='r',encoding='utf-8') as f:
            t2 = f.read()
            print "Same:" , t2 == txt 
            #print t2
def parse():
    global b
    import re
    items = []
    stop = False
    
    for i in range(500):
        pageno  = i+1
        filename = "files/%s.html" %pageno
        print "Loading page %d" %(pageno)
        txt = None
        with codecs.open("files/%s.html" %pageno, mode='r',encoding='utf-8') as f:
            txt = f.read()
        blocks  = re.findall("<!--start showResult-cell-->.*?<!--end showResult-cell-->", txt, re.S);
        for block in blocks:
            try:
                item = {}                                
                m = re.search(u'<p class="fb">(.*?)\s(.*?)<!--医院 科室 --></p>', block)
                if(m): 
                    item['hospital'] = m.group(1)
                    item['dept'] = m.group(2) 
                    #print "Hospital: %s Dept: %s" %(m.group(1), m.group(2))    
                m = re.search(u'>(.*?)<!--医生姓名--></a>(.*?)\s*<', block)
                if (m):
                    item['doctor'] = m.group(1)
                    item['title'] = m.group(2)
                m = re.search(u'(\d+)分钟（(\d+).00元）', block)
                if(m):
                    item['fee'] = m.group(2)
                    minutes = m.group(1)
                    item['fee_per_minute'] = int( int(m.group(2)) / float(minutes) )
                #print item
                m= re.search(u'<b class="orange">(\d+)</b>例;<b class="orange">(\d+)', block)
                if(m):
                    item['transactions'] = m.group(1)
                    item['reviews'] = m.group(2)
                                
                s="%s %s "  %(item['fee'], item['transactions'])
                item['subtotal'] = int(item['fee']) * int(item['transactions'])
                items.append(item)
            except:
                print block
                print traceback.format_exc()
            #print "%s %s DOCTOR %s HOSP %s TITLE %s"  %(item['fee'], item['transactions'], item['doctor'], item['title'], item['hospital'])
        
    print "\n\n### TOTAL ###"
    print "%s records found" %len(items)  
    
    from openpyxl.cell import get_column_letter
    from openpyxl import Workbook
    wb = Workbook()
    
    cols = {
            'hospital': 'A',
            'dept':'B',
            'doctor':'C',
            'title': 'D',
            'fee':'E',
            'fee_per_minute':'F',
            'transactions':'G',
            'reviews':'H', 
            'subtotal':'I'
            }
  
    ws = wb.worksheets[0]
    
    ws.title = u"电话咨询"
    row = 1
    for name, col in cols.items():
        ws.cell('%s%s' %(col, row)).value =  name
    for item in items:
        row =row + 1
        for name, col in cols.items():
            ws.cell('%s%s' %(col, row)).value = item[name]
    """
    for col_idx in xrange(1, 40):
        col = get_column_letter(col_idx)
        for row in xrange(1, 600):
            ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)
    """
    
    wb.save(filename = DataFile)
    #return results
         
#listAllCases()
if __name__ == '__main__':
    parse()
    #menus=harvestListOptions()
    #print menus
    #print listAllCases()
    if False:
        fields = getFields()
        for field in fields:
            print field['name'], field['type'], field['category']
            if('menu_name' in field):
                if field['menu_name'] in menus:
                    print field['menu_name'],'\t\t', menus[field['menu_name']]['options']
                else:
                    print '\t\t ERROR: missing menu %s' %field['menu_name']

htm = """
      <div class="mt5">
                                <div class="fl fb">收费标准：</div>
                                <div class="oh">
                                15分钟（150.00元）
&nbsp;
                                </div>
                            </div>
                            <div class="mt5">
                                <div class="fl fb">成功接听：</div>
                                <div class="oh">已成功接听<b class="orange">462</b>例;<b class="orange">414</b>条评价 <a href="http://400.haodf.com/haodf/zhengjinghao#successAnswer" target="_blank" class="blue">查看评价>> </a></div>
                            </div>
                            <div class="mt5 ml40 pl25">
                                <a href="http://zhengjinghao.haodf.com/payment/servicelist" target="_blank"><img src="http://i1.hdfimg.com/400/images/tel_doctor_now.jpg" alt="立刻预约" /></a>
                                拨打4006-606-120转3预约电话咨询
                            </div>
                        </div>
                    </div>
                    <!--end showResult-cell-->
                    <!--start showResult-cell-->
                    <div class="clearfix showResult-cell bb pb10 mt15">
                        <div class="showResult-cell-l fl">
                            <table cellspacing="0" cellpadding="0" border="0">
                                <tbody><tr>
                                    <td align="center" valign="middle" style="padding: 0px; border: 1px solid #DFDFDF; height: 110px; vertical-align: middle; width: 110px;">
                                    <a target="_blank" href="http://400.haodf.com/haodf/szspiderman"><img src="http://n1.hdfimg.com/cache/2010/1030/fa/9159d04f0f165b,110,110,1.jpg?6dc6"></a></td>
                                </tr>
                              </tbody>
                            </table>
                            <p class="tc mt5">
                            <!--大夫等级-->
                        <a href="http://400.haodf.com/haodf/szspiderman" target="_blank" class="green">麻晓鹏<!--医生姓名--></a>主任医师                          </p>
                        </div>                      <div class="showResult-cell-r fr">
                            <p class="fb">深圳儿童医院 小儿外科<!--医院 科室 --></p>
                            <div class="mt5">
                                <div class="fl fb">擅　　长：</div>
                                <div class="oh"><span>新生儿外科、普外科、小儿微创腔镜外科</span> <a href="#" class="blue showMore">查看详情>></a></div>
                            </div>


"""        